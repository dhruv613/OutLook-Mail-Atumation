import os
import time
import random
import threading
from datetime import datetime
from openpyxl import load_workbook

class RecipientStatus:
    # Status Implementation: "{STATUS}|{DATE}"
    # Example: "USED|16-01-2026"
    PROCESSING = "PROCESSING"
    USED = "USED"
    FAILED = "FAILED"

class RecipientExcelManager:
    _lock = threading.RLock()

    def __init__(self, file_path):
        self.file_path = file_path
        self.email_col = None
        self.status_col = None
        # self._last_processed_row = 2  # REMOVED: No longer used with Math Logic
        
        # Ensure strict file existence check
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"❌ Recipient Excel not found at: {self.file_path}")

        self._initialize_structure()

    # ----------------------------------------------------------------
    # 🔒 Helper: Safe Open / Save (Context Manager Approach preferred?)
    # Since we need very granular control, we'll use explicit methods
    # but wrap public methods in the lock.
    # ----------------------------------------------------------------

    def _open(self):
        """Helper to open workbook. Returns (wb, ws)."""
        # Load workbook with data_only=False to preserve formulas if any (though we manipulate values)
        # Using read_only=False because we intend to write.
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            return wb, ws
        except Exception as e:
            print(f"❌ Error opening Excel: {e}")
            raise

    def _safe_save(self, wb, retries=10):
        """Save workbook with retries for permission errors."""
        for i in range(retries):
            try:
                wb.save(self.file_path)
                return True
            except PermissionError:
                if i < retries - 1:
                    sleep_time = random.uniform(0.5, 2.0)
                    print(f"⚠️ Permission Error saving Recipient Excel. Retrying in {sleep_time:.2f}s...")
                    time.sleep(sleep_time)
                else:
                    print(f"❌ Failed to save Recipient Excel after {retries} retries: Permission Denied.")
                    raise
            except Exception as e:
                print(f"❌ Failed to save Recipient Excel: {e}")
                raise

    def _get_today_str(self):
        return datetime.now().strftime("%d-%m")

    # ----------------------------------------------------------------
    # 🚀 Core Functions
    # ----------------------------------------------------------------

    def _initialize_structure(self):
        """
        Detects Email/Status columns.
        Renames Status header to Today if needed.
        Does NOT scan rows (Lazy Reset).
        """
        with self._lock:
            wb, ws = self._open()
            try:
                # 1. Detect Columns
                today_str = self._get_today_str()
                
                found_email = None
                found_status = None
                status_header_val = None

                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=1, column=col).value
                    val_str = str(val).strip().lower() if val else ""
                    
                    if "email" in val_str:
                        found_email = col
                    elif val_str == "status" or val_str == today_str or (val and self._is_date_format(str(val))):
                        found_status = col
                        status_header_val = str(val).strip()

                # Set Email Column
                if found_email:
                    self.email_col = found_email
                else:
                    print("⚠️ 'Email' column not found. Defaulting to Column 1.")
                    self.email_col = 1

                # Set/Update Status Column
                if found_status:
                    self.status_col = found_status
                    # Rule: If header != today, Rename to today
                    if status_header_val != today_str:
                        print(f"🔄 New Day/Header Detected: Renaming '{status_header_val}' to '{today_str}'")
                        ws.cell(row=1, column=self.status_col).value = today_str
                        self._safe_save(wb)
                else:
                    # Create new Status Column
                    self.status_col = ws.max_column + 1
                    print(f"🆕 Creating Status Column at {self.status_col} with header '{today_str}'")
                    ws.cell(row=1, column=self.status_col).value = today_str
                    self._safe_save(wb)
                    
                # 2. Crash Recovery: Reset any stuck 'PROCESSING' rows (even from today)
                # If the script is initializing, any 'PROCESSING' status is stale from a crash.
                stuck_count = 0
                for row in range(2, ws.max_row + 1):
                    val = ws.cell(row=row, column=self.status_col).value
                    if val and RecipientStatus.PROCESSING in str(val).upper():
                        ws.cell(row=row, column=self.status_col).value = None # Reset to Available
                        stuck_count += 1
                
                if stuck_count > 0:
                    print(f"🔄 Checkpoint Recovery: Released {stuck_count} stuck '{RecipientStatus.PROCESSING}' rows.")
                    self._safe_save(wb)

            finally:
                wb.close()

    
    def get_batch_recipients(self, batch_size, sender_row_index):
        """
        Retrieves recipients using Stratified Sampling (Gap Formula).
        Formula: idx = (k * gap + sender_index) % total
        """
        with self._lock:
            wb, ws = self._open()
            recipients = []
            rows_to_update = []
            
            try:
                today_str = self._get_today_str()
                max_row = ws.max_row
                total_recipients = max_row - 1 # Exclude header
                
                if total_recipients <= 0:
                     return [], []

                gap = total_recipients // batch_size
                if gap < 1: gap = 1

                # Sender Index (0-based relative to sender list, but here we just need a unique seed)
                # sender_row_index comes from sender_excel (e.g., 2, 3, 4...)
                # Let's normalize it to 0-based for the math
                sender_seed = sender_row_index 
                
                # Shift offset in case of collisions
                shift = 0 
                
                # We need 'batch_size' valid recipients
                collected_count = 0
                
                # Safety loop to prevent infinite loops if all marked USED
                attempts = 0
                max_attempts = max(batch_size * 100, 2000) # Increased limit to handle lane congestion
                
                # Debug stats
                collisions = 0

                while collected_count < batch_size and attempts < max_attempts:
                    # Current virtual index 'i' in the batch
                    k = collected_count 
                    
                    # Calculate target index (0-based data index)
                    # Formula: (k * gap + sender_seed + shift) % total
                    target_data_idx = (k * gap + sender_seed + shift) % total_recipients
                    
                    # Convert to Excel Row (1-based header + 1-based index) -> Row 2 is data_idx 0
                    target_row = target_data_idx + 2
                    
                    # Check status
                    status_val = ws.cell(row=target_row, column=self.status_col).value
                    email_val = ws.cell(row=target_row, column=self.email_col).value
                    
                    if not email_val:
                        # Should not happen if total_recipients is correct, but safety
                        shift += 1
                        attempts += 1
                        continue

                    # Check availability
                    if self._is_status_available(status_val, today_str):
                         # Found valid!
                         recipients.append(str(email_val).strip())
                         rows_to_update.append(target_row)
                         collected_count += 1
                         # Reset shift for next item? local shift optimization
                    else:
                         # Collision (Used/Failed today) -> Try next
                         shift += 1
                    
                    attempts += 1
                
                if attempts >= max_attempts:
                    print(f"⚠️ Warning: High collision rate. Attempts: {attempts}/{max_attempts}. Retrieved {len(recipients)}/{batch_size}. Shift reached: {shift}")

                # --- FALLBACK: Linear Scan ---
                if len(recipients) < batch_size:
                    # print(f"⚠️ Gap search failed (Found {len(recipients)}/{batch_size}). Starting Linear Scan Fallback...")
                    
                    # Start scanning from the last successful hit or from row 2
                    # We scan the WHOLE file if needed until we fill the batch or hit end
                    
                    scan_limit = max_row
                    current_scan_row = 2
                    
                    while len(recipients) < batch_size and current_scan_row <= max_row:
                        # Skip if already in our current batch (rows_to_update check)
                        if current_scan_row in rows_to_update:
                             current_scan_row += 1
                             continue
                             
                        # Check status
                        status_val = ws.cell(row=current_scan_row, column=self.status_col).value
                        email_val = ws.cell(row=current_scan_row, column=self.email_col).value
                        
                        if not email_val:
                            current_scan_row += 1
                            continue

                        if self._is_status_available(status_val, today_str):
                              recipients.append(str(email_val).strip())
                              rows_to_update.append(current_scan_row)
                        
                        current_scan_row += 1
                        
                    if len(recipients) < batch_size:
                         print(f"❌ Linear Scan exhausted. Final count: {len(recipients)}/{batch_size}")

                # Note: We do NOT write PROCESSING status.
                
                # [CONCURRENCY FIX] Mark items as PROCESSING immediately to prevent collisions
                if rows_to_update:
                    full_status = f"{RecipientStatus.PROCESSING}|{today_str}"
                    for r in rows_to_update:
                        ws.cell(row=r, column=self.status_col).value = full_status
                    self._safe_save(wb)
                    # print(f"🔒 Locked {len(rows_to_update)} rows as '{RecipientStatus.PROCESSING}'")

                return recipients, rows_to_update

            finally:
                wb.close()

    def update_batch_status(self, rows, status):
        """
        Updates status for specific rows.
        Format: STATUS|DATE
        """
        if not rows:
            return

        with self._lock:
            wb, ws = self._open()
            try:
                today_str = self._get_today_str()
                full_status = f"{status}|{today_str}"
                
                for r in rows:
                    ws.cell(row=r, column=self.status_col).value = full_status
                
                self._safe_save(wb)
                # print(f"📝 Updated {len(rows)} rows to '{full_status}'.")
            
            except Exception as e:
                print(f"❌ Error updating recipient status: {e}")
            finally:
                wb.close()

    # ----------------------------------------------------------------
    # 🧠 Logic Helpers
    # ----------------------------------------------------------------

    def _is_status_available(self, status_val, today_str):
        """
        Determines if a row is available.
        Available if:
        - Status is None/Empty
        - Status is explicitly 'AVAILABLE' (custom)
        - Status contains a DIFFERENT date than today (Lazy Reset)
        
        NOT Available if:
        - Status contains TODAY's date (PROCESSING|Today, USED|Today, FAILED|Today)
        """
        if not status_val:
            return True
            
        s_val = str(status_val).strip()
        if not s_val:
            return True
            
        # Check for Date Suffix
        if "|" in s_val:
            parts = s_val.split("|")
            # format: STATUS|DATE
            if len(parts) >= 2:
                date_part = parts[-1]
                if date_part != today_str:
                    return True # Old date -> Available
                else:
                    return False # Today's status -> Not Available
        
        # If no date suffix found (e.g. legacy "USED"), 
        # it is treated as "Old" (Available) because we only write with dates now.
        return True

    def _is_date_format(self, text):
        # Rough check if text looks like a date/status header
        if "-" in text or "Status" in text:
            return True
        return False

    def get_used_count(self):
        """Count rows marked as USED (today or legacy)"""
        with self._lock:
            wb, ws = self._open()
            try:
                count = 0
                if not self.status_col:
                    return 0
                
                for row in range(2, ws.max_row + 1):
                    val = ws.cell(row=row, column=self.status_col).value
                    if val and str(val).strip().upper().startswith(RecipientStatus.USED):
                        count += 1
                return count
            finally:
                wb.close()
