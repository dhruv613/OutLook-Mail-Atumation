import os
import re
from datetime import datetime
from openpyxl import load_workbook
from automation.data.database_manager import DatabaseManager
from utils.config import SENDER_EXCEL_PATH, RECIPIENT_EXCEL_PATH

class SyncManager:
    def __init__(self):
        self.db = DatabaseManager()

    def _get_today_str_sender(self):
        return datetime.now().strftime("%d-%m-%Y")

    def _get_today_str_recipient(self):
        return datetime.now().strftime("%d-%m")

    # ---------------------------------------------------------
    # IMPORT: Excel -> SQLite
    # ---------------------------------------------------------
    def import_from_excel(self):
        print("📥 Starting Data Import (Excel -> SQLite)...")
        self._import_senders()
        self._import_recipients()
        print("✅ Data Import Complete.")

    def _import_senders(self):
        if not os.path.exists(SENDER_EXCEL_PATH):
            print(f"❌ Sender Excel missing: {SENDER_EXCEL_PATH}")
            return

        wb = load_workbook(SENDER_EXCEL_PATH) # Open Read-Only? No, we might update header
        ws = wb.active
        
        # 1. Detect Columns
        col_map = {}
        header_row = 1
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            s_val = str(val).strip().lower() if val else ""
            if "email" in s_val: col_map['email'] = col
            elif "password" in s_val: col_map['password'] = col
        
        if 'email' not in col_map:
            print(f"❌ Error: 'email' column not found in Sender Excel ({SENDER_EXCEL_PATH})")
            wb.close()
            return
        
        # 2. Daily Status Logic
        today_str = self._get_today_str_sender()
        status_col = None
        
        # Find status column
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            s_val = str(val).strip()
            # Check for Date Pattern or 'Status'
            if re.match(r"\d{1,2}-\d{1,2}-\d{4}", s_val) or s_val.lower() == "status":
                status_col = col
                break
        
        # If header date mismatch -> NEW DAY -> Reset logic
        is_new_day = False
        if status_col:
            header_val = str(ws.cell(row=header_row, column=status_col).value).strip()
            if header_val != today_str:
                print(f"🔄 Senders: New Day Detected ({header_val} != {today_str}). Resetting 'USED' status.")
                ws.cell(row=header_row, column=status_col).value = today_str 
                is_new_day = True
                
                # IMMEDIATE FIX: Reset statuses in the Excel file itself to prevent 'trap' state on crash
                print("   ...Cleaning Excel file statuses safely...")
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=status_col).value
                    if val and "USED" in str(val).upper():
                        ws.cell(row=r, column=status_col).value = None
                        
                wb.save(SENDER_EXCEL_PATH)
                print("   ...Excel file updated and saved with clean statuses.")
        else:
             print("⚠️ Status column not found in Sender Excel.")

        # 3. Read Data & Insert to DB
        senders_data = []
        
        # We assume data starts row 2
        for row in range(2, ws.max_row + 1):
            email = ws.cell(row=row, column=col_map.get('email')).value
            password = ws.cell(row=row, column=col_map.get('password')).value
            status = ws.cell(row=row, column=status_col).value if status_col else None
            
            if not email: continue
            
            email = str(email).strip()
            status = str(status).strip() if status else None
            
            # Apply Reset Logic
            if is_new_day:
                if status == "USED" or (status and "USED" in status):
                    status = None # Reset to Available
                # Keep BLOCKED, FAILED, NOT_LOGINED? 
                # Plan says: Keep BLOCKED, FAILED, NOT_LOGINED, NEED_PREMIUM.
                # Reset USED.
                pass
            
            # [NEW] Always reset FAILED status to allow retry on new run
            if status == "FAILED": 
                 status = None
            
            senders_data.append((email, password, status, 0, row))

        # Bulk Insert (Upsert)
        # We use INSERT OR REPLACE to update existing local DB cache if we want persistence, 
        # BUT for this session-based architecture, we might just TRUNCATE and Load?
        # Let's TRUNCATE first to be 100% synced with Excel state on startup.
        self.db.execute("DELETE FROM senders") 
        self.db.execute_many("""
            INSERT OR IGNORE INTO senders (email, password, status, rounds_completed, original_row)
            VALUES (?, ?, ?, ?, ?)
        """, senders_data)
        
        print(f"   --> Imported {len(senders_data)} Senders.")
        wb.close()

    def _import_recipients(self):
        if not os.path.exists(RECIPIENT_EXCEL_PATH):
             print("❌ Recipient file missing.")
             return

        wb = load_workbook(RECIPIENT_EXCEL_PATH)
        ws = wb.active
        
        # 1. Detect Columns
        email_col = None
        status_col = None
        today_str = self._get_today_str_recipient()
        
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            s_val = str(val).strip().lower() if val else ""
            if "email" in s_val: email_col = col
            elif s_val == "status" or s_val == today_str or (val and "-" in str(val)):
                status_col = col

        if not email_col:
            print(f"❌ Error: 'email' column not found in Recipient Excel ({RECIPIENT_EXCEL_PATH})")
            wb.close()
            return

        # Header Update Logic
        if status_col:
             header_val = str(ws.cell(row=1, column=status_col).value).strip()
             if header_val != today_str:
                 print(f"🔄 Recipients: New Day Detected. Updating header to {today_str}.")
                 ws.cell(row=1, column=status_col).value = today_str
                 
                 # IMMEDIATE FIX: Clean Excel statuses
                 print("   ...Cleaning Recipient Excel file statuses...")
                 for r in range(2, ws.max_row + 1):
                     ws.cell(row=r, column=status_col).value = None
                 
                 wb.save(RECIPIENT_EXCEL_PATH)


        recipients_data = []
        for row in range(2, ws.max_row + 1):
            email = ws.cell(row=row, column=email_col).value
            status = ws.cell(row=row, column=status_col).value if status_col else None
            
            if not email: continue
            
            email = str(email).strip()
            # Recipient Logic: "STATUS|DATE"
            # If date != today, treat as None (Available) during DB load
            
            db_status = None
            if status:
                s_status = str(status).strip()
                if "|" in s_status:
                    parts = s_status.split("|")
                    if parts[-1] == today_str:
                        db_status = parts[0] # Keep status (USED/FAILED)
                    else:
                        db_status = None # Reset
                else:
                    # Legacy status without date -> Reset
                    db_status = None 
            
            recipients_data.append((email, db_status, row))

        self.db.execute("DELETE FROM recipients")
        self.db.execute_many("""
            INSERT OR IGNORE INTO recipients (email, status, original_row)
            VALUES (?, ?, ?)
        """, recipients_data)
        
        print(f"   --> Imported {len(recipients_data)} Recipients.")
        wb.close()


    # ---------------------------------------------------------
    # EXPORT: SQLite -> Excel
    # ---------------------------------------------------------
    def export_to_excel(self):
        print("📤 Starting Data Export (SQLite -> Excel)...")
        self._export_senders()
        self._export_recipients()
        print("✅ Data Export Complete.")

    def _export_senders(self):
        # Read ALL from DB
        rows = self.db.fetch_all("SELECT original_row, status, rounds_completed FROM senders")
        
        wb = load_workbook(SENDER_EXCEL_PATH)
        ws = wb.active
        
        # Find Status Column again
        today_str = self._get_today_str_sender()
        status_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val and (str(val).strip() == today_str or "Status" in str(val)):
                status_col = col
                break
        
        if not status_col:
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col).value = today_str

        # Update
        for item in rows:
            r_idx = item['original_row']
            stat = item['status']
            # If status is PENDING and rounds > 0, we can format it? 
            # Logic says: PENDING:{rounds}
            
            rounds = item['rounds_completed']
            if stat == "PENDING" and rounds > 0:
                final_val = f"PENDING:{rounds}"
            else:
                final_val = stat
            
            ws.cell(row=r_idx, column=status_col).value = final_val

        wb.save(SENDER_EXCEL_PATH)
        wb.close()

    def _export_recipients(self):
        # Read Modified only? Or all? Best to read only those with status
        rows = self.db.fetch_all("SELECT original_row, status FROM recipients WHERE status IS NOT NULL")
        
        wb = load_workbook(RECIPIENT_EXCEL_PATH)
        ws = wb.active
        
        today_str = self._get_today_str_recipient()
        status_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val and str(val).strip() == today_str:
                status_col = col
                break
        
        if not status_col:
             status_col = ws.max_column + 1
             ws.cell(row=1, column=status_col).value = today_str

        for item in rows:
            r_idx = item['original_row']
            stat = item['status']
            # Format: STATUS|DATE
            full_stat = f"{stat}|{today_str}"
            ws.cell(row=r_idx, column=status_col).value = full_stat

        wb.save(RECIPIENT_EXCEL_PATH)
        wb.close()
