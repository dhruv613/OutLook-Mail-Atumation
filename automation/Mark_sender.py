from openpyxl import load_workbook


class SenderExcelManager:
    def __init__(self, sender_excel_path):
        self.sender_excel_path = sender_excel_path
        self.email_col = None
        self.password_col = None
        self.status_col = None

        self.current_row = 2  # start after header

        self._detect_columns()

    # --------------------------------------------------------------------
    # PRIVATE: Detect email/password/status columns
    # --------------------------------------------------------------------
    def _detect_columns(self):
        wb = load_workbook(self.sender_excel_path)
        ws = wb.active

        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col).value).strip().lower()

            if header == "email":
                self.email_col = col
            elif header == "password":
                self.password_col = col
            elif header == "status":
                self.status_col = col

        wb.close()

        if not self.email_col or not self.password_col:
            raise ValueError("❌ Excel me 'Email', 'Password' columns nahi mile.")

    # --------------------------------------------------------------------
    # PUBLIC: Get next available Email / Password where status != USED
    # --------------------------------------------------------------------
    def get_next_sender(self):
        wb = load_workbook(self.sender_excel_path)
        ws = wb.active

        for row in range(self.current_row, ws.max_row + 1):

            email = ws.cell(row=row, column=self.email_col).value
            password = ws.cell(row=row, column=self.password_col).value
            status = ws.cell(row=row, column=self.status_col).value if self.status_col else None

            # Rule:
            #   ✔ email + password must exist
            #   ✔ status must NOT be USED
            if email and password and status != "USED":
                self.current_row = row
                wb.close()
                return email, password, row

        wb.close()
        return None, None, None

    # --------------------------------------------------------------------
    # PUBLIC: Mark row as USED
    # --------------------------------------------------------------------
    def mark_sender_used(self, row):
        wb = load_workbook(self.sender_excel_path)
        ws = wb.active

        if not self.status_col:
            self.status_col = ws.max_column + 1
            ws.cell(row=1, column=self.status_col).value = "Status"

        ws.cell(row=row, column=self.status_col).value = "USED"
        wb.save(self.sender_excel_path)
        wb.close()
