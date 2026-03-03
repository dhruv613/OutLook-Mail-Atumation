import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from automation.data.database_manager import DatabaseManager
from utils.config import SENDER_EXCEL_PATH

def debug_sender_counts():
    print(f"🔍 DEBUGGING SENDER COUNTS")
    print(f"📂 Excel Path: {SENDER_EXCEL_PATH}")
    
    if not os.path.exists(SENDER_EXCEL_PATH):
        print("❌ Excel file not found!")
        return

    # 1. ANALYZE EXCEL DIRECTLY
    print("\n--- 1. EXCEL FILE ANALYSIS ---")
    try:
        wb = load_workbook(SENDER_EXCEL_PATH, data_only=True)
        ws = wb.active
        
        # Find headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers[col] = str(val).strip()
        
        print(f"   Headers found: {headers}")
        
        # Identify Status Column (looking for Date or 'Status')
        status_col = None
        for col, name in headers.items():
            if "Status" in name or "-" in name: # Simple check for date-like header
                status_col = col
                print(f"   ✅ Identified Status Column: '{name}' (Col {col})")
                break
        
        if not status_col:
            print("   ⚠️ Could not identify Status column in Excel.")
        else:
            # Count Statuses
            status_counts = {}
            row_details = []
            
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=status_col).value
                email = ws.cell(row=r, column=1).value # Assuming Email is Col 1, adjust if needed
                
                if val:
                    s_val = str(val).strip()
                    status_counts[s_val] = status_counts.get(s_val, 0) + 1
                    
                    if "USED" in s_val.upper():
                        row_details.append(f"Row {r}: {email} -> '{s_val}'")
            
            print(f"   📊 Excel Status Counts (Raw):")
            for k, v in status_counts.items():
                print(f"      - '{k}': {v}")
                
            print(f"\n   📝 Details for 'USED' entries in Excel:")
            for d in row_details[:10]: # Print first 10
                print(f"      {d}")
            if len(row_details) > 10:
                print(f"      ... and {len(row_details) - 10} more.")

    except Exception as e:
        print(f"❌ Error reading Excel: {e}")

    # 2. ANALYZE DATABASE
    print("\n--- 2. SQLITE DB ANALYSIS ---")
    try:
        db = DatabaseManager()
        # Check raw counts
        rows = db.fetch_all("SELECT status, COUNT(*) as cnt FROM senders GROUP BY status")
        
        print(f"   📊 DB Status Counts:")
        for r in rows:
            print(f"      - '{r['status']}': {r['cnt']}")
            
        # Check 'Used' specific
        used_rows = db.fetch_all("SELECT email, status FROM senders WHERE status LIKE 'USED%' OR status = 'USED'")
        print(f"\n   📝 DB Entries matching 'USED%': {len(used_rows)}")
        for r in used_rows[:10]:
            print(f"      - {r['email']}: '{r['status']}'")

    except Exception as e:
         print(f"❌ Error reading DB: {e}")

if __name__ == "__main__":
    debug_sender_counts()
