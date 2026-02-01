import unittest
import os
from openpyxl import Workbook
import random

# Mocking the Manager to avoid full dependency chain if needed, 
# or importing directly if path allows. 
# We'll import the class but we need to setup a dummy excel.

from automation.excel.recipient_excel import RecipientExcelManager

class TestCollisionFix(unittest.TestCase):
    FILE_NAME = "test_recipients.xlsx"
    BATCH_SIZE = 30
    TOTAL_ROWS = 1000
    
    def setUp(self):
        # Create a dummy excel
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Email"
        ws.cell(row=1, column=2).value = "19-01" # Today's header
        
        # Fill with dummy emails
        for i in range(2, self.TOTAL_ROWS + 2):
            ws.cell(row=i, column=1).value = f"user{i}@example.com"
            # Randomly mark some as USED to simulate fragmentation
            if random.random() < 0.5: # 50% used
                ws.cell(row=i, column=2).value = "USED|19-01"
                
        wb.save(self.FILE_NAME)
        self.mgr = RecipientExcelManager(self.FILE_NAME)

    def tearDown(self):
        try:
            os.remove(self.FILE_NAME)
        except:
            pass

    def test_retrieval_under_load(self):
        print("\nTesting retrieval with 50% collision rate...")
        
        # Try to retrieve a batch
        recipients, rows = self.mgr.get_batch_recipients(self.BATCH_SIZE, sender_row_index=5)
        
        print(f"Requested: {self.BATCH_SIZE}")
        print(f"Retrieved: {len(recipients)}")
        
        self.assertEqual(len(recipients), self.BATCH_SIZE, "Should retrieve full batch even with high collisions")

if __name__ == '__main__':
    unittest.main()
